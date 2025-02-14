# reusable utility functions
from selenium import webdriver
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.edge.service import Service as EdgeService
from selenium.webdriver.firefox.service import Service as FirefoxService
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import openpyxl
import configparser
import os
from logging_config import logging

def initialize_driver(browser, config):
    """Initializes the WebDriver."""
    try:
        browser = browser.lower()
        drivers = {
            "chrome": webdriver.Chrome,
            "edge": webdriver.Edge,
            "firefox": webdriver.Firefox,
        }
        if browser not in drivers:
            raise ValueError(f"Unsupported browser: {browser}")
        driver_path = config.get("PATHS", f"{browser}driver")
        service = ChromeService(executable_path=driver_path) if browser == "chrome" else None
        driver = drivers[browser](service=service)
        driver.maximize_window()
        return driver
    except Exception as e:
        logging.error(f"Error initializing driver: {e}")
        raise

def scroll_into_view(driver, *element):
    """Scrolls an element into view."""
    try:
        driver.execute_script("arguments[0].scrollIntoView();", driver.find_element(*element))
    except Exception as e:
        logging.warning(f"Could not scroll to element: {e}")

def select_dropdown_option(driver, dropdown_locator, value, option_location):
    """Selects an option from a dropdown."""
    try:
        dropdown = WebDriverWait(driver, 10).until(EC.presence_of_element_located(dropdown_locator))
        options = dropdown.find_elements(*option_location)
        for option in options:
            if option.text.lower() == value.lower():
                option.click()
                return
        logging.warning(f"Option {value} not found in dropdown.")
    except TimeoutException:
        logging.error("Dropdown not found or not clickable.")

def read_excel_data(sheet_name, config):
    """Reads data from an Excel file."""
    try:
        excel_file_path = config.get("PATHS", "excelfile")
        work_book = openpyxl.load_workbook(excel_file_path)
        sheet = work_book[sheet_name]
        columns_number = sheet.max_column
        rows_number = sheet.max_row
        data_frame = {}
        for col in range(1, columns_number + 1):
            cell_value = []
            for row in range(2, rows_number + 1):
                cell_value.append(sheet.cell(row, col).value)
            data_frame[sheet.cell(1, col).value] = cell_value
        return data_frame
    except FileNotFoundError:
        logging.error(f"Excel file not found: {excel_file_path}")
        return {}

def quit_browser(driver):
    """Closes the browser."""
    driver.quit()

def get_project_root():
    """Returns the project root directory."""
    current_dir = os.path.dirname(os.path.abspath(__file__))
    while not os.path.exists(os.path.join(current_dir, "README.md")):
        current_dir = os.path.dirname(current_dir)
    return current_dir

def read_config():
    """Reads the configuration from config.ini."""
    config = configparser.ConfigParser()
    project_root = get_project_root()
    config_path = os.path.join(project_root, "config", "config.ini")
    config.read(config_path)
    return config